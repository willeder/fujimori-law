#!/usr/bin/env python3
"""
docs/data の CSV（入金予定）と xlsx（依頼者基本情報・受任案件管理・和解処理）から
client-mock 用の mockCases / mockPaymentRecords / mockCreditors を生成する。

使い方（リポジトリルート）:
  . .venv-xlsx/bin/activate
  python3 scripts/generate_mock_from_docs.py
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "docs/data/CSVファイル.csv"
XLSX_PATH = ROOT / "docs/data/受任案件管理_マージ済_20260330.xlsx"
OUT_CASES = ROOT / "client-mock/src/data/mockCasesFromDocs.ts"
OUT_PAYMENTS = ROOT / "client-mock/src/data/mockPaymentsFromDocs.ts"
OUT_CREDITORS = ROOT / "client-mock/src/data/mockCreditorsFromDocs.ts"

# メインシート上の出現順で、依頼者シート・CSV の両方に存在する ID のうち末尾 N 件を採用
SELECT_CASE_COUNT = 30

ADJUSTMENT_ALLOWED = {"任意整理", "自己破産", "個人再生"}
GENDER_ALLOWED = {"男", "女"}
MARITAL_ALLOWED = {"既婚", "未婚", "離婚"}

CREDITOR_STATUS_ALLOWED = {
    "受任通知発送待ち",
    "受任通知発送済",
    "債権調査中",
    "和解提案中",
    "和解済",
    "弁済中",
    "完済",
}
CREDITOR_STATUS_MAP = {
    "債権調査票待ち": "債権調査中",
    "債権調査中": "債権調査中",
    "和解提案書発送待ち": "和解提案中",
    "和解提案書発送済": "和解提案中",
    "和解提案中": "和解提案中",
    "一部受任通知発送済": "受任通知発送済",
    "全社受任通知発送済": "受任通知発送済",
    "全和解済_支払中": "弁済中",
    "一部和解済_支払中": "弁済中",
}


def to_iso_date(val) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s or s == "########":
        return None
    s = s.replace("/", "-")
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    return None


def ym_cell(val) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m")
    if isinstance(val, date):
        return val.strftime("%Y-%m")
    s = str(val).strip()
    if re.match(r"^\d{4}-\d{2}$", s):
        return s
    return None


def to_num(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if not s or s == "########":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def to_int(val) -> int | None:
    n = to_num(val)
    if n is None:
        return None
    return int(n)


def to_str(val) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    s = str(val).strip()
    return s or None


def csv_date(s: str) -> str | None:
    s = (s or "").strip()
    if not s:
        return None
    if "/" in s:
        parts = s.split("/")
        if len(parts) == 3:
            y, m, d = parts
            return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
    return None


def pad_row(row: tuple) -> list:
    r = list(row or ())
    if len(r) < 56:
        r.extend([None] * (56 - len(r)))
    return r


def row_triple(row: tuple) -> tuple[float, float, float] | None:
    r = pad_row(row)
    L = len(row or ())
    if L >= 28:
        a, b, c = r[12], r[13], r[14]
    elif L == 15:
        a, b, c = r[11], r[12], r[13]
    else:
        return None
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) and isinstance(c, (int, float)):
        return (float(a), float(b), float(c))
    return None


def normalize_creditor_status(raw: str | None) -> str:
    if not raw:
        return "和解提案中"
    s = str(raw).strip()
    if s in CREDITOR_STATUS_ALLOWED:
        return s
    return CREDITOR_STATUS_MAP.get(s, "和解提案中")


def load_main_index(ws) -> tuple[dict[str, tuple], dict[str, int]]:
    hdr = [str(c) if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(hdr) if h}
    by_id: dict[str, tuple] = {}
    i_id = idx["ID"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[i_id] is None:
            continue
        eid = str(row[i_id]).strip()
        by_id[eid] = row
    return by_id, idx


def load_client_rows(ws) -> dict[str, dict]:
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    by_id: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[1] is None:
            continue
        eid = str(row[1]).strip()
        cells = list(row) + [None] * max(0, len(hdr) - len(row))
        by_id[eid] = {hdr[i]: cells[i] for i in range(len(hdr))}
    return by_id


def discover_selected_external_ids() -> list[str]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws_c = wb["依頼者　基本情報"]
    client_ids: set[str] = set()
    for row in ws_c.iter_rows(min_row=2, values_only=True):
        if row and row[1]:
            client_ids.add(str(row[1]).strip())

    ws_m = wb["受任案件管理_マージ済_20260330"]
    hdr = list(next(ws_m.iter_rows(min_row=1, max_row=1, values_only=True)))
    midx = {h: i for i, h in enumerate(hdr) if h}
    i_id = midx["ID"]

    csv_ids: set[str] = set()
    with open(CSV_PATH, "r", encoding="cp932", newline="") as f:
        cr = csv.reader(f)
        next(cr)
        for row in cr:
            if len(row) > 1 and row[1].strip():
                csv_ids.add(row[1].strip())

    order: list[str] = []
    for row in ws_m.iter_rows(min_row=2, values_only=True):
        if not row or row[i_id] is None:
            continue
        eid = str(row[i_id]).strip()
        if eid in client_ids and eid in csv_ids:
            order.append(eid)
    wb.close()
    if len(order) < SELECT_CASE_COUNT:
        raise SystemExit(f"need {SELECT_CASE_COUNT} intersection IDs, got {len(order)}")
    return order[-SELECT_CASE_COUNT:]


def load_settlement_rows() -> list[tuple | None]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["和解処理"]
    rows: list[tuple | None] = [None]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append(tuple(row) if row else tuple())
    wb.close()
    return rows


def build_settlement_blocks(rows: list[tuple | None]) -> list[tuple[int, int, tuple[float, float, float]]]:
    blocks: list[tuple[int, int, tuple[float, float, float]]] = []
    cur: tuple[float, float, float] | None = None
    start = 3
    max_r = len(rows) - 1
    for ri in range(3, max_r + 1):
        row = rows[ri]
        t = row_triple(row) if row else None
        if t is None:
            continue
        if cur is None:
            cur, start = t, ri
        elif t != cur:
            blocks.append((start, ri - 1, cur))
            cur, start = t, ri
    if cur is not None:
        blocks.append((start, max_r, cur))
    return blocks


def count_creditor_table_rows(rows: list[tuple | None], s: int, e: int) -> int:
    n = 0
    for ri in range(s, e + 1):
        row = rows[ri]
        if not row:
            continue
        if len(row) >= 28 and row[15] and isinstance(row[15], str):
            name = str(row[15]).strip()
            if name and not name.startswith("★") and name != "債権者":
                n += 1
    return n


@dataclass
class SettlementBlock:
    start: int
    end: int
    triple: tuple[float, float, float]

    @property
    def mid_row(self) -> float:
        return (self.start + self.end) / 2


def assign_blocks_to_cases(
    rows: list[tuple | None],
    blocks: list[tuple[int, int, tuple[float, float, float]]],
    case_specs: list[tuple[str, int, float | None, int | None, int]],
) -> dict[str, SettlementBlock | None]:
    """
    case_specs: (eid, main_excel_row, decl, creditor_count, _)
    """
    enriched: list[SettlementBlock] = []
    for s, e, t in blocks:
        cc = count_creditor_table_rows(rows, s, e)
        if cc <= 0:
            continue
        enriched.append(SettlementBlock(start=s, end=e, triple=t))

    used: set[int] = set()
    out: dict[str, SettlementBlock | None] = {}
    for eid, main_row, decl, creditor_count, _ in sorted(case_specs, key=lambda x: x[1]):
        cc_i = int(creditor_count) if creditor_count is not None else 0
        if cc_i <= 0:
            out[eid] = None
            continue

        best_i: int | None = None
        best_score: float = -10**18
        decl_f = float(decl) if decl is not None else None

        for i, b in enumerate(enriched):
            if i in used:
                continue
            if cc_i > 0 and count_creditor_table_rows(rows, b.start, b.end) != cc_i:
                continue
            if decl_f is None:
                score = 0.0
            else:
                err = abs(b.triple[0] - decl_f)
                if err < 0.5:
                    score = 1_000_000.0
                else:
                    score = 500_000.0 - err
            score -= abs(b.mid_row - main_row) * 0.01
            if score > best_score:
                best_score = score
                best_i = i

        if best_i is None:
            out[eid] = None
        else:
            used.add(best_i)
            out[eid] = enriched[best_i]
    return out


def negotiation_partner_cell(val) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        if val == 0:
            return None
        if val == int(val):
            return str(int(val))
        return None
    s = str(val).strip()
    if s in ("0", "0.0"):
        return None
    return s or None


def emit_creditor_ts_object(rows: list[tuple | None], ri: int, case_id: int, creditor_id: int) -> list[str]:
    row = pad_row(rows[ri])
    name = to_str(row[15]) or "（名称不明）"
    st = normalize_creditor_status(to_str(row[19]))

    def jsn(x):
        return json.dumps(x, ensure_ascii=False)

    lines: list[str] = []
    lines.append("  {")
    lines.append(f"    id: {creditor_id},")
    lines.append(f"    caseId: {case_id},")
    lines.append(f"    creditorName: {jsn(name)},")
    lines.append(f"    negotiationPartner: {jsn(negotiation_partner_cell(row[16]))},")
    lines.append(f"    declaredAmount: {json.dumps(to_int(row[17]))},")
    lines.append(f"    debtAmount: {json.dumps(to_int(row[26]))},")
    lines.append(f"    expectedSettlement: {json.dumps(to_int(row[18]))},")
    lines.append(f"    status: {jsn(st)},")
    lines.append(f"    nextProcessDate: {jsn(to_iso_date(row[21]))},")
    lines.append(f"    acceptanceNoticeSentDate: {jsn(to_iso_date(row[22]))},")
    lines.append(f"    debtInquiryArrivalDate: {jsn(to_iso_date(row[23]))},")
    lines.append(f"    customerCode: {jsn(to_str(row[24]))},")
    lines.append(f"    contractDate: {jsn(to_iso_date(row[25]))},")
    lines.append(f"    settlementProposalDate: {jsn(to_iso_date(row[28]) if len(row) > 28 else None)},")
    lines.append(f"    responseStatus: {jsn(to_str(row[30]) if len(row) > 30 else None)},")
    lines.append(f"    settlementDate: {jsn(to_iso_date(row[31]) if len(row) > 31 else None)},")
    lines.append(f"    settlementAmount: {json.dumps(to_int(row[32]) if len(row) > 32 else None)},")
    lines.append(f"    settlementDebtAmount: {json.dumps(to_int(row[33]) if len(row) > 33 else None)},")
    lines.append(f"    settlementContentComment: {jsn(to_str(row[34]) if len(row) > 34 else None)},")
    lines.append(f"    paymentStartMonth: {jsn(ym_cell(row[39]) if len(row) > 39 else None)},")
    pd = to_int(row[40]) if len(row) > 40 else None
    lines.append(f"    paymentDay: {json.dumps(pd)},")
    lines.append(f"    paymentCount: {json.dumps(to_int(row[41]) if len(row) > 41 else None)},")
    lines.append(f"    firstPaymentAmount: {json.dumps(to_int(row[42]) if len(row) > 42 else None)},")
    lines.append(f"    subsequentPaymentAmount: {json.dumps(to_int(row[43]) if len(row) > 43 else None)},")
    lines.append(f"    finalPaymentAmount: {json.dumps(to_int(row[44]) if len(row) > 44 else None)},")
    lines.append(f"    finalPaymentMonth: {jsn(ym_cell(row[45]) if len(row) > 45 else None)},")
    fi = to_str(row[46]) if len(row) > 46 else None
    if fi in ("0", "0.0", None):
        fi = None
    lines.append(f"    futureInterest: {jsn(fi or 'なし')},")
    lines.append(f"    bankName: {jsn(to_str(row[47]) if len(row) > 47 else None)},")
    lines.append(f"    branchName: {jsn(to_str(row[49]) if len(row) > 49 else None)},")
    lines.append(f"    accountType: {jsn(to_str(row[51]) if len(row) > 51 else None)},")
    an = row[52] if len(row) > 52 else None
    if isinstance(an, float) and an == int(an):
        an = str(int(an))
    lines.append(f"    accountNumber: {jsn(to_str(an))},")
    lines.append(f"    accountHolder: {jsn(to_str(row[53]) if len(row) > 53 else None)},")
    lines.append("  },")
    return lines


def stub_creditor_lines(case_id: int, creditor_id: int, idx: int, decl: int | None, cc: int) -> list[str]:
    part = max(1, cc)
    guess = int(decl / part) if decl else None

    def jsn(x):
        return json.dumps(x, ensure_ascii=False)

    return [
        "  {",
        f"    id: {creditor_id},",
        f"    caseId: {case_id},",
        f"    creditorName: {json.dumps(f'（和解シート未突合 {idx}/{cc}）', ensure_ascii=False)},",
        "    negotiationPartner: null,",
        f"    declaredAmount: {json.dumps(guess)},",
        "    debtAmount: null,",
        "    expectedSettlement: null,",
        '    status: "和解提案中",',
        "    nextProcessDate: null,",
        "    acceptanceNoticeSentDate: null,",
        "    debtInquiryArrivalDate: null,",
        "    customerCode: null,",
        "    contractDate: null,",
        "    settlementProposalDate: null,",
        "    responseStatus: null,",
        "    settlementDate: null,",
        "    settlementAmount: null,",
        "    settlementDebtAmount: null,",
        "    settlementContentComment: null,",
        "    paymentStartMonth: null,",
        "    paymentDay: null,",
        "    paymentCount: null,",
        "    firstPaymentAmount: null,",
        "    subsequentPaymentAmount: null,",
        "    finalPaymentAmount: null,",
        "    finalPaymentMonth: null,",
        f"    futureInterest: {jsn('なし')},",
        "    bankName: null,",
        "    branchName: null,",
        "    accountType: null,",
        "    accountNumber: null,",
        "    accountHolder: null,",
        "  },",
    ]


def normalize_adjustment(s: str | None) -> str | None:
    if not s:
        return None
    s = str(s).strip()
    if s in ADJUSTMENT_ALLOWED:
        return s
    return "任意整理"


def normalize_gender(s: str | None) -> str | None:
    if not s:
        return None
    s = str(s).strip()
    if s in GENDER_ALLOWED:
        return s
    return None


def normalize_marital(s: str | None) -> str | None:
    if not s:
        return None
    s = str(s).strip()
    if s in MARITAL_ALLOWED:
        return s
    return "未婚"


def elapsed_days(accept: str | None, ref: date) -> int | None:
    if not accept:
        return None
    try:
        d0 = date.fromisoformat(accept)
        return (ref - d0).days
    except ValueError:
        return None


def emit_cases(
    ref_date: date,
    selected_ids: list[str],
    main_by_id: dict[str, tuple],
    midx: dict[str, int],
    client_by_id: dict[str, dict],
) -> str:
    lines: list[str] = []
    lines.append("/**")
    lines.append(" * 自動生成: scripts/generate_mock_from_docs.py")
    lines.append(" * 元: docs/data/受任案件管理_マージ済_20260330.xlsx / docs/data/CSVファイル.csv")
    lines.append(" */")
    lines.append("import type { Case } from '../types/case'")
    lines.append("")
    lines.append("export const mockCasesFromDocs: Case[] = [")

    for num, eid in enumerate(selected_ids, start=1):
        c = client_by_id.get(eid)
        m = main_by_id.get(eid)
        if not c or not m:
            raise SystemExit(f"missing data for {eid}")

        def mi(name: str):
            j = midx.get(name)
            return m[j] if j is not None and j < len(m) else None

        name = to_str(c.get("名前")) or "（氏名未設定）"
        furigana = to_str(c.get("フリガナ"))
        phone = to_str(c.get("電話番号"))
        email = to_str(c.get("メールアドレス"))
        line_url = to_str(c.get("LINE@ URL"))
        caution = to_str(c.get("要注意ランク"))
        if caution and caution not in ("A", "B", "C"):
            caution = None

        list_cat = to_str(c.get("リスト区分"))
        list_reg = to_iso_date(c.get("リスト登録日"))

        birth = to_iso_date(c.get("生年月日"))
        age = to_int(c.get("年齢"))
        gender = normalize_gender(to_str(c.get("性別")))
        marital = normalize_marital(to_str(c.get("結婚")))
        children = to_str(c.get("子供"))
        cohab = to_str(c.get("同居"))
        confid = to_str(c.get("内密先"))
        emer = to_str(c.get("緊急連絡先"))
        emer_rel = to_str(c.get("関係(緊急)"))
        prev_addr = to_str(c.get("旧住所"))
        monthly = to_int(c.get("月収(手取)"))
        pay_day_raw = c.get("給与日")
        pay_day = (
            str(int(pay_day_raw))
            if isinstance(pay_day_raw, float) and pay_day_raw == int(pay_day_raw)
            else to_str(pay_day_raw)
        )
        payroll = to_str(c.get("給与口座"))
        emp_name = to_str(c.get("勤務先名"))
        emp_type = to_str(c.get("勤務形態"))
        emp_contact = to_str(c.get("勤務先連絡先"))
        emp_addr = to_str(c.get("勤務先住所"))
        other_off = to_str(c.get("他事務所相談"))
        delay = to_str(c.get("遅れ"))
        bike = to_str(c.get("自転車"))
        pension = to_str(c.get("年金"))
        record_no = to_int(c.get("レコード番号"))
        corr_req = to_str(c.get("対応要否"))
        corr_hr = to_str(c.get("対応時間"))

        acc_date = to_iso_date(mi("受任日"))
        interview = to_str(mi("面談担当"))
        appt = to_str(mi("アポ担当"))
        follow = to_str(mi("後確担当"))
        jud = to_str(mi("担当司法書士"))
        adj = normalize_adjustment(to_str(mi("債務整理区分")))
        rank_raw = to_str(mi("受任ランク"))
        rank = rank_raw if rank_raw in ("A", "B", "C") else None

        debt_cc = to_int(mi("債権社数"))
        decl = to_int(mi("申告債務額"))
        total_d = to_int(mi("債務額総額"))
        pre_p = to_int(mi("依頼 前 返済額"))
        post_p = to_int(mi("依頼 後 返済額"))

        st = to_str(mi("受任後ステータス"))
        prop = to_iso_date(mi("和解提案予定日"))
        scount = to_int(mi("和解弁済総数"))
        pcount = to_int(mi("予定弁済総数"))
        post_c = to_int(mi("和解後代弁社数"))
        plan_agent = to_int(mi("予定代弁社数"))
        all_sent = to_iso_date(mi("全和解書送付日"))

        nf = to_int(mi("通常報酬"))
        inst = to_int(mi("報酬分割回数"))
        agent_p = to_str(mi("弁済代行"))
        plan_fee_tot = to_int(mi("予定弁済報酬総額"))
        uncoll = to_int(mi("報酬未回収額"))

        fp_date = to_iso_date(mi("初回入金予定日"))
        fp_amt = to_int(mi("初回入金額"))
        monthly_pay_day = to_str(mi("毎月入金日"))
        base_amt = to_int(mi("基本入金額"))
        next_pay = to_iso_date(mi("次回入金日"))
        cum_pay = to_int(mi("累)入金金額"))
        cum_plan = to_int(mi("累)入金予定額"))
        cum_fee = to_int(mi("累)報酬充当額"))
        cum_plan_fee = to_int(mi("累)報酬充当予定額"))
        cum_pool = to_int(mi("累)ﾌﾟｰﾙ充当額"))
        cum_rep = to_int(mi("累)弁済充当額"))
        tot_minus = to_int(mi("総額-ﾌﾟｰﾙ-累弁済"))

        rem_d = to_iso_date(mi("リマインド日"))
        rem_t = to_str(mi("リマインド時間"))
        next_resp = to_iso_date(mi("次回対応日"))

        created = to_iso_date(mi("作成日時"))
        updated = to_iso_date(mi("更新日時"))
        created_by = to_str(mi("作成者"))
        updated_by = to_str(mi("更新者"))

        elap = elapsed_days(acc_date, ref_date)

        c_promo = to_iso_date(mi("C受任昇格日"))
        memo1 = to_str(mi("面談時備考１"))
        memo2 = to_str(mi("面談時備考２"))
        inc_memo = to_str(mi("収支メモ"))
        within10 = to_str(mi("10日以内"))

        def jsn(x):
            return json.dumps(x, ensure_ascii=False)

        lines.append("  {")
        lines.append(f"    id: {num},")
        lines.append(f"    clientBasicInfo: {{")
        lines.append(f"      name: {jsn(name)},")
        lines.append(f"      furigana: {jsn(furigana)},")
        lines.append(f"      phone: {jsn(phone)},")
        if line_url:
            lines.append(f"      lineUrl: {jsn(line_url)},")
        lines.append(f"      email: {jsn(email)},")
        lines.append(f"      prefecture: {jsn(to_str(c.get('都道府県')))},")
        lines.append(f"      address: {jsn(to_str(c.get('住所')))},")
        lines.append(f"      birthDate: {jsn(birth)},")
        lines.append(f"      age: {json.dumps(age)},")
        lines.append(f"      gender: {json.dumps(gender)},")
        lines.append(f"      maritalStatus: {json.dumps(marital)},")
        lines.append(f"      children: {jsn(children)},")
        lines.append(f"      residenceType: {jsn(to_str(c.get('居住形態')))},")
        lines.append(f"      rent: {json.dumps(to_int(c.get('家賃')))},")
        lines.append(f"      monthlyIncome: {json.dumps(monthly)},")
        lines.append(f"      payDay: {jsn(pay_day)},")
        lines.append(f"      employmentType: {jsn(emp_type)},")
        lines.append(f"      cautionRank: {json.dumps(caution)},")
        lines.append(f"      recordNumber: {json.dumps(record_no)},")
        lines.append(f"      correspondenceRequired: {jsn(corr_req)},")
        lines.append(f"      correspondenceHours: {jsn(corr_hr)},")
        lines.append(f"      cohabitation: {jsn(cohab)},")
        lines.append(f"      confidentialContact: {jsn(confid)},")
        lines.append(f"      emergencyContact: {jsn(emer)},")
        lines.append(f"      emergencyContactRelation: {jsn(emer_rel)},")
        lines.append(f"      previousAddress: {jsn(prev_addr)},")
        lines.append(f"      payrollAccount: {jsn(payroll)},")
        lines.append(f"      employerName: {jsn(emp_name)},")
        lines.append(f"      employerContact: {jsn(emp_contact)},")
        lines.append(f"      employerAddress: {jsn(emp_addr)},")
        lines.append(f"      otherOfficeConsultation: {jsn(other_off)},")
        lines.append(f"      paymentDelay: {jsn(delay)},")
        lines.append(f"      bicycleNote: {jsn(bike)},")
        lines.append(f"      pension: {jsn(pension)},")
        lines.append("    },")
        lines.append("    appointmentInfo: {")
        lines.append(f"      appointmentStaff: {jsn(appt)},")
        lines.append(f"      followUpStaff: {jsn(follow)},")
        lines.append(f"      interviewStaff: {jsn(interview)},")
        lines.append(f"      judicialScrivener: {jsn(jud)},")
        lines.append(f"      debtAdjustmentType: {json.dumps(adj)},")
        lines.append(f"      acceptanceRank: {json.dumps(rank)},")
        lines.append(f"      acceptanceDate: {jsn(acc_date)},")
        lines.append(f"      elapsedDays: {json.dumps(elap)},")
        lines.append(f"      cAcceptancePromotionDate: {jsn(c_promo)},")
        lines.append(f"      interviewMemo1: {jsn(memo1)},")
        lines.append(f"      interviewMemo2: {jsn(memo2)},")
        lines.append(f"      incomeExpenseMemo: {jsn(inc_memo)},")
        lines.append("    },")
        lines.append("    debtInfo: {")
        lines.append(f"      creditorCount: {json.dumps(debt_cc)},")
        lines.append(f"      declaredDebtAmount: {json.dumps(decl)},")
        lines.append(f"      totalDebtAmount: {json.dumps(total_d)},")
        lines.append(f"      preRequestPayment: {json.dumps(pre_p)},")
        lines.append(f"      postRequestPayment: {json.dumps(post_p)},")
        lines.append("    },")
        lines.append("    settlementInfo: {")
        lines.append(f"      status: {jsn(st)},")
        lines.append(f"      proposalDate: {jsn(prop)},")
        lines.append(f"      settlementCount: {json.dumps(scount)},")
        lines.append(f"      postSettlementPaymentCount: {json.dumps(post_c)},")
        lines.append(f"      plannedPaymentCount: {json.dumps(pcount)},")
        lines.append(f"      plannedAgentCount: {json.dumps(plan_agent)},")
        lines.append(f"      allSettlementDocSentDate: {jsn(all_sent)},")
        lines.append("    },")
        lines.append("    feeInfo: {")
        lines.append(f"      normalFee: {json.dumps(nf)},")
        lines.append(f"      officeFee: {json.dumps(nf)},")
        lines.append(f"      installmentCount: {json.dumps(inst)},")
        lines.append(f"      agentPayment: {jsn(agent_p)},")
        lines.append(f"      plannedPaymentFeeTotal: {json.dumps(plan_fee_tot)},")
        lines.append(f"      uncollectedFee: {json.dumps(uncoll)},")
        lines.append("    },")
        lines.append("    paymentInfo: {")
        lines.append(f"      firstPaymentDate: {jsn(fp_date)},")
        lines.append(f"      firstPaymentWithinTenDays: {jsn(within10)},")
        lines.append(f"      firstPaymentAmount: {json.dumps(fp_amt)},")
        lines.append(f"      monthlyPaymentDay: {jsn(monthly_pay_day)},")
        lines.append(f"      basePaymentAmount: {json.dumps(base_amt)},")
        lines.append(f"      nextPaymentDate: {jsn(next_pay)},")
        lines.append(f"      cumulativePaymentAmount: {json.dumps(cum_pay)},")
        lines.append(f"      cumulativePlannedPayment: {json.dumps(cum_plan)},")
        lines.append(f"      cumulativeFeeAllocation: {json.dumps(cum_fee)},")
        lines.append(f"      cumulativePlannedFeeAllocation: {json.dumps(cum_plan_fee)},")
        lines.append(f"      cumulativePoolAllocation: {json.dumps(cum_pool)},")
        lines.append(f"      cumulativeRepaymentAllocation: {json.dumps(cum_rep)},")
        lines.append(f"      totalMinusPoolMinusRepayment: {json.dumps(tot_minus)},")
        lines.append("      vAccountBranch: null,")
        lines.append("      vAccountNumber: null,")
        lines.append("    },")
        lines.append("    reminderInfo: {")
        lines.append(f"      reminderDate: {jsn(rem_d)},")
        lines.append(f"      reminderTime: {jsn(rem_t)},")
        lines.append(f"      nextResponseDate: {jsn(next_resp)},")
        lines.append("      responseTime: null,")
        lines.append("    },")
        lines.append("    metadata: {")
        lines.append(f"      createdAt: {jsn(created)},")
        lines.append(f"      updatedAt: {jsn(updated)},")
        lines.append(f"      createdBy: {jsn(created_by)},")
        lines.append(f"      updatedBy: {jsn(updated_by)},")
        lines.append(f"      externalId: {jsn(eid)},")
        lines.append(f"      listCategory: {jsn(list_cat)},")
        lines.append(f"      listRegisteredDate: {jsn(list_reg)},")
        lines.append("      acceptanceDocs: '',")
        lines.append("    },")
        lines.append("  },")

    lines.append("]")
    lines.append("")
    return "\n".join(lines)


def emit_payments(selected_ids: list[str]) -> str:
    id_to_case_num = {eid: i + 1 for i, eid in enumerate(selected_ids)}

    rows_by_case: dict[int, list[list[str]]] = {i: [] for i in id_to_case_num.values()}
    with open(CSV_PATH, "r", encoding="cp932", newline="") as f:
        r = csv.reader(f)
        next(r)
        for row in r:
            if len(row) < 22:
                continue
            eid = str(row[1]).strip()
            if eid not in id_to_case_num:
                continue
            rows_by_case[id_to_case_num[eid]].append(row)

    lines: list[str] = []
    lines.append("/**")
    lines.append(" * 自動生成: scripts/generate_mock_from_docs.py")
    lines.append(" * 元: docs/data/CSVファイル.csv（入金予定履歴相当）")
    lines.append(" */")
    lines.append("import type { PaymentRecord } from '../types/case'")
    lines.append("")
    lines.append("export const mockPaymentRecordsFromDocs: PaymentRecord[] = [")

    pay_id = 1
    for eid in selected_ids:
        case_num = id_to_case_num[eid]

        block = rows_by_case[case_num]
        block.sort(key=lambda rr: (rr[4] or ""))
        for row in block:
            planned_date = csv_date(row[4])
            planned_amt = to_int(row[5])
            pfee = to_int(row[6])
            pagent = to_int(row[7])
            ppool = to_int(row[8])
            rep_cnt = to_int(row[9])
            hand = to_int(row[10])
            prepay = to_int(row[11])
            act_date = csv_date(row[13])
            act_amt = to_int(row[14])
            afee = to_int(row[15])
            aagent = to_int(row[16])
            apool = to_int(row[17])
            arepay = to_int(row[21])
            cum_pool = to_int(row[26]) if len(row) > 26 else None

            lines.append("  {")
            lines.append(f"    id: {pay_id},")
            lines.append(f"    caseId: {case_num},")
            lines.append("    creditorId: null,")
            lines.append("    creditorInstallmentIndex: null,")
            lines.append(f"    plannedDate: {json.dumps(planned_date)},")
            lines.append(f"    plannedAmount: {json.dumps(planned_amt)},")
            lines.append(f"    plannedFeeAllocation: {json.dumps(pfee)},")
            lines.append(f"    plannedAgentFeeAllocation: {json.dumps(pagent)},")
            lines.append(f"    plannedPoolAllocation: {json.dumps(ppool)},")
            lines.append(f"    plannedRepaymentAllocation: {json.dumps(prepay)},")
            lines.append(f"    actualDate: {json.dumps(act_date)},")
            lines.append(f"    actualAmount: {json.dumps(act_amt)},")
            lines.append(f"    actualFeeAllocation: {json.dumps(afee)},")
            lines.append(f"    actualAgentFeeAllocation: {json.dumps(aagent)},")
            lines.append(f"    actualPoolAllocation: {json.dumps(apool)},")
            lines.append(f"    actualRepaymentAllocation: {json.dumps(arepay)},")
            lines.append(f"    handlingFee: {json.dumps(hand)},")
            lines.append(f"    repaymentCount: {json.dumps(rep_cnt)},")
            lines.append(f"    cumulativePool: {json.dumps(cum_pool)},")
            lines.append("  },")
            pay_id += 1

    lines.append("]")
    lines.append("")
    return "\n".join(lines)


def emit_creditors(
    selected_ids: list[str],
    rows: list[tuple | None],
    eid_to_block: dict[str, SettlementBlock | None],
    main_by_id: dict[str, tuple],
    midx: dict[str, int],
) -> str:
    lines: list[str] = []
    lines.append("/**")
    lines.append(" * 自動生成: scripts/generate_mock_from_docs.py")
    lines.append(" * 元: docs/data/受任案件管理_マージ済_20260330.xlsx の和解処理シート")
    lines.append(" */")
    lines.append("import type { Creditor } from '../types/case'")
    lines.append("")
    lines.append("export const mockCreditorsFromDocs: Creditor[] = [")

    creditor_id = 1
    for case_num, eid in enumerate(selected_ids, start=1):
        m = main_by_id[eid]

        def mi(name: str):
            j = midx.get(name)
            return m[j] if j is not None and j < len(m) else None

        decl = to_int(mi("申告債務額"))
        cc = to_int(mi("債権社数")) or 0
        blk = eid_to_block.get(eid)

        if blk:
            wrote = 0
            for ri in range(blk.start, blk.end + 1):
                row = rows[ri]
                if not row or len(row) < 28:
                    continue
                if not row[15] or not isinstance(row[15], str):
                    continue
                name = str(row[15]).strip()
                if not name or name.startswith("★") or name == "債権者":
                    continue
                lines.extend(emit_creditor_ts_object(rows, ri, case_num, creditor_id))
                creditor_id += 1
                wrote += 1
            if wrote == 0 and cc > 0:
                for k in range(cc):
                    lines.extend(stub_creditor_lines(case_num, creditor_id, k + 1, decl, cc))
                    creditor_id += 1
        elif cc and cc > 0:
            for k in range(cc):
                lines.extend(stub_creditor_lines(case_num, creditor_id, k + 1, decl, cc))
                creditor_id += 1

    lines.append("]")
    lines.append("")
    return "\n".join(lines)


def main():
    ref = date(2026, 4, 12)
    selected_ids = discover_selected_external_ids()

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws_main = wb["受任案件管理_マージ済_20260330"]
    ws_client = wb["依頼者　基本情報"]
    main_by_id, midx = load_main_index(ws_main)
    client_by_id = load_client_rows(ws_client)

    hdr = [str(c) if c is not None else "" for c in next(ws_main.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(hdr) if h}
    i_id = idx["ID"]

    main_row_by_eid: dict[str, int] = {}
    for ri, row in enumerate(ws_main.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[i_id] is None:
            continue
        eid = str(row[i_id]).strip()
        main_row_by_eid[eid] = ri
    wb.close()

    case_specs: list[tuple[str, int, float | None, int | None, int]] = []
    for eid in selected_ids:
        m = main_by_id[eid]
        mr = main_row_by_eid[eid]

        def mi(name: str):
            j = midx.get(name)
            return m[j] if j is not None and j < len(m) else None

        decl = to_num(mi("申告債務額"))
        cc = to_int(mi("債権社数"))
        case_specs.append((eid, mr, decl, cc, 0))

    settlement_rows = load_settlement_rows()
    raw_blocks = build_settlement_blocks(settlement_rows)
    eid_to_block = assign_blocks_to_cases(settlement_rows, raw_blocks, case_specs)

    OUT_CASES.parent.mkdir(parents=True, exist_ok=True)
    OUT_CASES.write_text(emit_cases(ref, selected_ids, main_by_id, midx, client_by_id), encoding="utf-8")
    OUT_PAYMENTS.write_text(emit_payments(selected_ids), encoding="utf-8")
    OUT_CREDITORS.write_text(
        emit_creditors(selected_ids, settlement_rows, eid_to_block, main_by_id, midx),
        encoding="utf-8",
    )
    print("Wrote", OUT_CASES)
    print("Wrote", OUT_PAYMENTS)
    print("Wrote", OUT_CREDITORS)
    print("Selected", len(selected_ids), "cases:", selected_ids[0], "…", selected_ids[-1])


if __name__ == "__main__":
    main()
